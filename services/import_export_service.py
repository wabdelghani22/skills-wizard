import os
import json
from openpyxl import Workbook
from docx import Document
from db import crud


class ImportExportService:
    def __init__(self, projet_id):
        self.projet_id = projet_id

    def export_to_excel(self, filepath: str):
        """Exporte toutes les données liées au projet dans un fichier Excel structuré."""
        wb = Workbook()

        # --- Feuille 1 : Hiérarchie ---
        ws1 = wb.active
        ws1.title = "Hiérarchie"
        ws1.append(["Famille", "Sous-Famille", "Métier"])

        for row in get_all_jobs_with_hierarchy_by_projet(self.projet_id):
            ws1.append(row)

        # --- Feuille 2 : Référentiel de compétences ---
        ws2 = wb.create_sheet("Référentiel")
        ws2.append(["Compétence", "Type", "Description"])
        for comp in get_all_competences_by_projet(self.projet_id):
            comp_id, nom, type_, description = comp
            ws2.append([nom, type_, description or ""])

        # --- Feuille 3 : Fiches de poste ---
        ws3 = wb.create_sheet("Fiches de poste")
        ws3.append(["Poste", "Activité", "Compétence", "Niveau"])
        jobs = get_all_jobs_with_hierarchy_by_projet(self.projet_id)
        job_comp_dict = get_job_competences_detailed_by_projet(self.projet_id)

        for _, _, job_name in jobs:
            activities = get_job_activities_by_projet(job_name, self.projet_id)
            competences = job_comp_dict.get(job_name, [])

            for act in activities or [""]:
                for comp in competences or [{"nom": "", "niveau": ""}]:
                    ws3.append([job_name, act, comp["nom"], comp.get("niveau", "")])

        # --- Feuille 4 : Dictionnaire de compétences ---
        ws4 = wb.create_sheet("Dictionnaire")
        ws4.append(["Compétence", "Niveau", "Définition du niveau"])
        for comp in get_all_competences_by_projet(self.projet_id):
            comp_id, nom, _, _ = comp
            niveau_defs = get_dictionnaire_definitions_by_competence_id(comp_id)
            for niveau, definition in niveau_defs.items():
                ws4.append([nom, niveau, definition or ""])

        wb.save(filepath)

    def export_to_json(self, filepath: str, data: dict):
        """Exporte un dictionnaire Python vers un fichier JSON."""
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

    def import_from_json(self, filepath: str):
        """Importe un fichier JSON brut (retourne un dict)."""
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)

    def extract_text_from_docx(self, filepath: str) -> str:
        """Extrait le texte brut d’un fichier Word."""
        doc = Document(filepath)
        return "\n".join([p.text.strip() for p in doc.paragraphs if p.text.strip()])

    def import_job_files_from_folder(self, folder_path: str, extensions=[".docx"]) -> dict:
        """Extrait les textes des fichiers Word dans un dossier."""
        data = {}
        for filename in os.listdir(folder_path):
            if any(filename.endswith(ext) for ext in extensions):
                filepath = os.path.join(folder_path, filename)
                data[filename] = self.extract_text_from_docx(filepath)
        return data
